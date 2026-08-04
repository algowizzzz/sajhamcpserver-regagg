#!/usr/bin/env python3
"""
Generate the PO-facing PM pack: a forward-looking 6-sprint roadmap (planning/)
and a parallel execution/status pack incl. the full test-case inventory
(execution/). One dataset drives both, so they can never drift.

Outputs:
  docs/regagg/planning/  sprint_plan.csv · epics.csv · capacity.csv
                         Regulatory_Aggregator_Roadmap.xlsx · ROADMAP.md
  docs/regagg/execution/ story_status.csv · test_cases.csv
                         Execution_Status.xlsx · STATUS_REPORT.md
"""

from __future__ import annotations

import csv
import subprocess
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
PLAN = REPO / "docs/regagg/planning"
EXEC = REPO / "docs/regagg/execution"
PLAN.mkdir(parents=True, exist_ok=True)
EXEC.mkdir(parents=True, exist_ok=True)

# ── epics ────────────────────────────────────────────────────────────────────
EPICS = [
    ("E1", "Foundation & Configuration Framework",
     "Config-driven skeleton: regulator-as-config YAML model, corpus database schema, storage layout, CI gates. Adding regulator #31 must never require code."),
    ("E2", "Versioning & Governed Storage",
     "The trust core: one current copy per document, append-only archive, crash-safe atomic override protocol, full provenance (source URL, hash, run attribution, raw original)."),
    ("E3", "Collection Engine",
     "Three stateless connector engines (sitemap-diff / RSS / API), fetch & HTML/PDF→markdown normalization, PDF harvesting, dedup, pagination, crawl hardening."),
    ("E4", "Source Verification & Regulator Onboarding",
     "No source is trusted until the verification harness passes it. Onboard all 30 regulators with recorded test fixtures and provable completeness (expected-inventory reconciliation)."),
    ("E5", "Orchestration, Operations & Data Quality",
     "Fleet fan-out with failure isolation, humane run statuses, live progress, daily delta pollers, manual interjection lane, backups, bot-block fallbacks."),
    ("E6", "Deterministic Enrichment & Knowledge Graph",
     "No-LLM enrichment: reference-number grammars per regulator, citation mining into a supersedes/references graph, expected-inventory completeness checks."),
    ("E7", "Analyst & Operator Dashboard",
     "Five-page UI: Coverage tree with coverage-%, Corpus browser, file Explorer, Changes/impact feed with diffs, Collection-runs monitor with live counters."),
    ("E8", "MCP Tool Surface & Security",
     "Twelve stateless reg_* MCP tools (content + index planes), API-key lifecycle, authentication required for execution, per-key tool allowlists enforced."),
    ("E9", "Agentic Integration",
     "Plug the corpus into the existing agent stack: markdown projection in the agent's consumption layout, scoped MCP connectivity, chatbot guide, agent skills, citation eval."),
    ("E10", "Future Enhancements (Backlog)",
     "Post-MVP: govinfo full-text connector, simple-mode overview UX, Playwright fetch, Postgres, more inventories, digests, production hardening, LLM enrichment."),
]

# ── stories ──────────────────────────────────────────────────────────────────
# (id, epic, sprint, title, description, acceptance criteria, points, role,
#  status, evidence/note)
D, NS, PR = "Done", "Not Started", "Partial"
S = [
 # Sprint 1 — Foundation
 ("REG-101","E1",1,"Project scaffolding & CI test harness",
  "Stand up the regagg module inside the MCP server repo with an isolated pytest harness (in-memory DB + temp storage) runnable on every commit.",
  "Module imports cleanly inside host server; pytest suite runs offline in <10s; CI-safe fixtures for DB and storage exist",
  3,"Backend",D,"tests/regagg/conftest.py; suite runs 92 tests in ~3s"),
 ("REG-102","E1",1,"Regulator configuration schema & loader",
  "Strict pydantic model for per-regulator YAML (connector type, sources, fetch method, rate limits, doc-type rules, backfill cutoff) with a fail-fast loader.",
  "Invalid/unknown fields rejected at load; connector↔sources cross-validated; filename must match declared id; loader collects all errors for reporting",
  5,"Backend",D,"config_models.py, config_loader.py; foundation gate 30/30"),
 ("REG-103","E1",1,"Author 30 regulator configs + controlled taxonomy",
  "One YAML per regulator across CA/US/EU-UK/APAC/INTL plus the controlled tag taxonomy and global settings files.",
  "30 configs parse and validate; taxonomy doc-types match code constants; every source starts verified:false",
  3,"Backend",D,"config/regulators/*.yaml (30) + _taxonomy + _settings"),
 ("REG-104","E1",1,"Corpus database schema (9 tables)",
  "SQLAlchemy models on the shared base for documents, versions, tags, edges, pending-edges, seen-URLs, runs, watermarks, regulators; matching hand-authored Postgres DDL.",
  "Tables auto-create on both SQLite and Postgres; one-current-per-doc expressible; FKs and CHECK constraints enforced; DDL file shipped",
  5,"Backend",D,"regagg/models.py; db/scripts/postgresql/003_regagg_schema.sql"),
 ("REG-105","E1",1,"Corpus storage layout over storage abstraction",
  "Normalized folder layout (current/archive/staging/_state per regulator) implemented over the host's local/S3 storage backend with artifact read/write helpers.",
  "Layout identical for all regulators; copy/delete tree operations work on local and S3 semantics; meta.json round-trips",
  3,"Full-stack",D,"corpus_storage.py"),
 ("REG-106","E1",1,"Foundation verification gate",
  "A no-network CI gate proving configs parse, taxonomy is in sync, and all corpus tables register/create.",
  "Single command exits non-zero on any foundation defect; run in CI on every commit",
  2,"Backend",D,"scripts/regagg_verify_foundation.py — GREEN"),
 ("REG-107","E4",1,"Source verification harness",
  "Automated verifier for every candidate source: reachability, content-type, parseability (incl. sitemap indexes), freshness; produces a per-regulator report.",
  "Pass/warn/fail per source with reasons; sitemap-index recognized; report renders to markdown; harness is the only authority for verified:true",
  5,"AI",D,"verify_sources.py; 29/30 regulators verified live"),
 ("REG-108","E4",1,"Fixture recorder & per-regulator parse tests",
  "Record each regulator's real source payloads once and run every connector against its recording in CI.",
  "Fixtures recorded for all reachable regulators; parametrized test per regulator passes offline; recordings size-capped",
  3,"AI",D,"regagg_record_fixtures.py; tests/fixtures/ 29/30 (9.6MB)"),
 ("REG-109","E1",1,"BA: source map validation & regulator list sign-off",
  "Validate the proposed 30-regulator list and candidate sources against business priorities; obtain sign-off.",
  "Signed-off regulator list; discrepancies (dead feeds, blocked sites) logged as onboarding risks",
  2,"BA",D,"Source map verified; AMF block + 5 dead feeds logged and later remediated"),
 ("PM-S1","E1",1,"PM: sprint ceremonies, RAID log, stakeholder reporting",
  "Recurring per-sprint PM allocation (~25%): planning, standups, review/retro, risk and dependency management.",
  "Ceremonies held; RAID log current; sprint report issued",
  0,"PM",D,"Recurring"),

 # Sprint 2 — Engine core
 ("REG-201","E2",2,"Atomic versioning protocol (override→archive)",
  "Six-step crash-safe protocol: stage new version, record intent, copy current to timestamped archive, promote staging, finalize document row.",
  "New doc creates v1; changed content archives prior version and increments version; unchanged content is a no-op; document identity stable across revisions",
  8,"Backend",D,"versioning.py; test_versioning.py"),
 ("REG-202","E2",2,"Crash-recovery reconciliation & invariant audit",
  "Deterministic repair job that rolls half-applied updates forward, cleans orphan staging, and audits invariants (exactly one current per doc; archive append-only).",
  "After any induced crash, reconcile restores a clean state with zero committed-data loss; violations are reported, never silently fixed twice",
  5,"Backend",D,"versioning.reconcile(); /integrity endpoint OK on 5,969 docs"),
 ("REG-203","E2",2,"Chaos test suite for the versioning path",
  "Automated tests that kill the pipeline after every protocol step and assert full recovery.",
  "Crash injected at each of steps 1–5; invariants hold post-reconcile in all cases; reconcile idempotent",
  3,"Backend",D,"test_versioning.py chaos parametrization (10 tests)"),
 ("REG-204","E3",2,"Sitemap-diff connector",
  "Detect new/changed URLs from sitemap.xml (incl. index recursion) plus listing pages, with lastmod fast-path skipping and include/exclude filters.",
  "Sitemap+listing union deduplicated; known URLs with unchanged lastmod skipped; doc-type mapped from URL rules",
  5,"AI",D,"connectors.SitemapDiffConnector; fixture + live tests"),
 ("REG-205","E3",2,"RSS/Atom connector",
  "Poll 1..n feeds per regulator with GUID→URL dedup, malformed-feed isolation, and backfill-cutoff filtering.",
  "Duplicate GUIDs collapse; items older than cutoff ignored; one bad feed never aborts others",
  3,"AI",D,"connectors.RssConnector; cutoff regression test"),
 ("REG-206","E3",2,"API connector & Federal Register adapter",
  "Provider-adapter API connector; Federal Register: agency filters, explicit field selection, pagination to cutoff, sanctioned full-text URL, abstract capture.",
  "Paginates until backfill cutoff; emits reference numbers for dedup; requests raw_text_url+abstract explicitly (not in default fields)",
  5,"AI",D,"connectors.ApiConnector; build_api_url; 500 docs/run"),
 ("REG-207","E3",2,"Fetch & normalization layer",
  "Polite fetcher (UA, per-domain rate limit) converting HTML→markdown (boilerplate stripped) and PDF→markdown, detecting PDFs by magic bytes, flagging scanned files.",
  "HTML preserves headings/tables/links; .pdf URLs serving HTML are treated as HTML (magic-byte truth); scanned PDFs flagged ocr:true; content sha256 computed",
  5,"Full-stack",D,"fetch.py; test_fetch.py incl. block-page regression"),
 ("REG-208","E2",2,"BA: provenance & audit acceptance walkthrough",
  "Define and verify the governance chain a compliance reviewer needs: source URL, final URL, hash, timestamps, run attribution, raw original, version lineage.",
  "Walkthrough from rendered doc back to original bytes and run manifest demonstrated; gaps (post-redirect URL) closed",
  2,"BA",D,"Provenance tab; meta.json fields; 0 docs missing source_url"),
 ("PM-S2","E5",2,"PM: sprint ceremonies & reporting","Recurring PM allocation.","As S1",0,"PM",D,"Recurring"),

 # Sprint 3 — Pipeline, ops core, onboarding wave 1
 ("REG-301","E3",3,"End-to-end ingestion pipeline",
  "Wire detect→fetch→version→enrich→record per regulator: run manifests, seen-URL registry, per-URL error isolation, idempotent reruns.",
  "One bad document never fails a run; rerun same day is a no-op for unchanged content; manifest lists every URL with outcome; run row persisted",
  8,"Backend",D,"pipeline.run_regulator; OSFI e2e test incl. forced update→archive"),
 ("REG-302","E3",3,"Policy-PDF harvesting",
  "Extract same-domain PDF links from ingested HTML pages and ingest them as first-class policy documents (source_kind=policy_pdf), with per-page and per-run caps.",
  "PDFs enter the same versioning path; source_kind recorded; caps prevent runaway; 3/page, 40/run defaults",
  3,"AI",D,"772 policy PDFs in corpus; harvest test"),
 ("REG-303","E3",3,"Meta-source deduplication",
  "When the meta-source (Federal Register) lists a document an agency feed already provided, keep the agency copy canonical and skip the duplicate by reference number.",
  "Duplicate skipped and counted (deduped metric); no double-counting across fedreg/agency",
  2,"AI",D,"meta_source flag; dedup test"),
 ("REG-304","E5",3,"Fleet orchestrator",
  "Fan out one isolated run per active regulator; support rerun of all/subset/one for a chosen date; expose reconcile as an operation.",
  "One regulator's hard failure never blocks others; rerun records operator identity; reruns are idempotent",
  5,"Backend",D,"orchestrator.py; isolation + rerun tests"),
 ("REG-305","E5",3,"Run status semantics & live progress",
  "Humane statuses (failed only when nothing landed or >20% errored) plus mid-run counter flushes so the UI can show live progress.",
  "Partial-error runs show success + error count; counters update at least every 10 documents during a run",
  3,"Backend",D,"events.finalize(); _flush_run; Today's-delta banner"),
 ("REG-306","E4",3,"Pilot regulators end-to-end (OSFI, FRB, FedReg)",
  "Prove all three connector types live against the pilot trio, including a forced content change correctly archiving v1.",
  "All pilots ingest live; update archives prior version; fixtures recorded",
  3,"AI",D,"Live runs; OSFI 833→1,196 docs"),
 ("REG-307","E4",3,"Onboarding wave 1 — fifteen fast regulators",
  "Verify sources and run first live ingestion for the non-giant regulators across all regions.",
  "Each has ≥1 verified source and a passing live run; failures triaged with per-URL evidence",
  5,"AI",D,"18 regs at 100% of advertised coverage"),
 ("REG-308","E3",3,"Crawl hardening",
  "Protect the fleet from pathological sources: sitemap byte cap, sitemap-index child cap, giants-last ordering with per-giant caps, tuned timeouts.",
  "A monster sitemap cannot hang a run (logged truncation); giants capped per run; hardening covered by tests",
  3,"Full-stack",D,"8MB/30-child caps after HKMA 3.8MB hang"),
 ("REG-309","E5",3,"Bot-block guard & sanctioned fallbacks",
  "Detect anti-scraping block pages and fall back to official API abstract, else an honest metadata stub — provenance labels content_source; never store block-page garbage; never evade.",
  "Block page never stored as content; fallback chain full-text→abstract→stub; content_source recorded; zero recurring error storms",
  2,"Full-stack",D,"FedReg: 64 abstracts + 436 stubs, 0 errors"),
 ("REG-310","E7",3,"BA: dashboard information architecture & status vocabulary",
  "Define analyst/operator journeys, coverage semantics (fresh/stale/attention), and plain-language status labels for the UI.",
  "IA for 5 pages agreed; status color/meaning table documented",
  2,"BA",D,"Mockups reviewed; humane-status change driven by PO feedback"),
 ("PM-S3","E5",3,"PM: sprint ceremonies & reporting","Recurring PM allocation.","As S1",0,"PM",D,"Recurring"),

 # Sprint 4 — Dashboard core, enrichment, wave 2
 ("REG-401","E7",4,"Dashboard data API",
  "Read endpoints backing every page: coverage tree with roll-ups, corpus browse with facets, changes feed, run overview with daily delta, version diff, inventory, jailed file explorer.",
  "All endpoints stateless and paginated where relevant; institution filters applied in SQL before LIMIT; fs endpoints jailed to corpus root",
  5,"Backend",D,"queries_ui.py + admin.py; endpoint tests incl. traversal + filter-order regressions"),
 ("REG-402","E7",4,"Coverage page",
  "Region→institution expandable tree: web/PDF counts, new-in-window, last run, freshness pill, coverage-% vs source-advertised, per-row Run trigger.",
  "Coverage % shown per institution; staleness flagged per config threshold; Run button triggers a real audited run",
  5,"Full-stack",D,"/ui Coverage; auto-refresh during active runs"),
 ("REG-403","E7",4,"Corpus browser page",
  "Dedicated deep-linkable page: continent/institution/file-type/doc-type/status/date/search filters with live facet counts, title+reference+excerpt rows, source links, Published+Added columns, pagination.",
  "All filters server-side; facet counts reflect scope; row opens document drawer; 50/page with prev/next",
  5,"Full-stack",D,"#corpus; excerpt layer exposed the FedReg block issue"),
 ("REG-404","E7",4,"Collection-runs page",
  "Operations view: today's delta banner with pass-rate headline and failing-run digest, live active-run progress (5s polling), recent runs, daily delta history.",
  "Pass-rate = fetched/(fetched+errors); live counters tick during runs; every run links to its manifest",
  3,"Full-stack",D,"#run; verified against live crawls"),
 ("REG-405","E6",4,"Reference-number grammar & citation mining",
  "Regex grammars for regulator identifiers (OSFI B/E-series, CAR/LAR chapters, SR letters, OCC bulletins, NI, APS, PS/CP/SS, RTS/ITS, BCBS…) extracting a document's own reference and its citations, with sentence-scoped supersede detection.",
  "Own-reference extracted from titles; citations classified supersedes vs references without cross-sentence bleed; adding a grammar is a table entry",
  5,"AI",D,"rules.py; pattern + sentence-scope tests"),
 ("REG-406","E6",4,"Cross-reference knowledge graph",
  "Resolve citations against the corpus into typed edges with confidence; supersedes edges flip target status; unresolved citations persist and retry nightly.",
  "Edges deduplicated; supersedes updates target to superseded; pending edges resolve when the target arrives",
  3,"AI",D,"176 edges + 756 pending on live corpus"),
 ("REG-407","E6",4,"Expected-inventory reconciliation (completeness proof)",
  "Per-regulator expected inventories (seeded: OSFI CAR/LAR/B/E series) reconciled against the corpus by reference then title; per-series present/missing report and API.",
  "Answers 'do we hold every chapter?' with named gaps; OSFI reconciles 30/30 after gap-fill",
  3,"AI",D,"/inventory/osfi = 30/30; _inventories/osfi.yaml"),
 ("REG-408","E4",4,"Onboarding wave 2 — giants & source remediation",
  "Onboard the large/broken regulators: capped depth-slicing for giants; fix dead feeds (FDIC), connector switches (HKMA + 5 RSS-dead regs → listing/sitemap), sitemap-index support (CSA).",
  "29/30 regulators verified and collecting; each fix re-verified by the harness; AMF escalated (not evaded)",
  3,"AI",D,"eiopa/finra/iosco/mas/osc/hkma/fdic/csa all remediated"),
 ("REG-409","E4",4,"Targeted gap-fill runs",
  "URL-regex-scoped ingest (e.g. /en/guidance) so specific sections can be completed without full-site crawls; used to complete OSFI's guideline chapters.",
  "Scoped run ingests only matching URLs; OSFI guidance section filled (363 docs) driving inventory to complete",
  2,"Backend",D,"--include flag; OSFI 12/30→30/30"),
 ("REG-410","E7",4,"BA: UAT round 1 (coverage/corpus/runs)",
  "Structured UAT of the first three pages against analyst/operator journeys; log and triage defects.",
  "UAT script executed; defects logged with severity; blocking defects fixed before sprint close",
  2,"BA",PR,"Continuous PO review happened in-build (several defects caught: filters, statuses); formal scripted UAT pending"),
 ("PM-S4","E5",4,"PM: sprint ceremonies & reporting","Recurring PM allocation.","As S1",0,"PM",D,"Recurring"),

 # Sprint 5 — UI completion, MCP, security, manual lane
 ("REG-501","E7",5,"Changes / impact feed",
  "Analyst delta feed: NEW/REVISED/SUPERSEDED/DEADLINE cards grouped by day, filtered by region, institution, file type, change-type chips, and explicit date range — all server-side.",
  "Filters share coverage-tree dimensions; counts remain pre-kind-filter for tile toggles; feed matches DB deltas exactly",
  3,"Full-stack",D,"#chg; filter-before-LIMIT regression test"),
 ("REG-502","E7",5,"Version diff viewer",
  "One-click unified diff between a document's current and previous archived version with added/removed counts, from the Changes feed and drawer.",
  "Diff renders color-coded; single-version docs report 'no diff' cleanly",
  2,"Full-stack",D,"/diff endpoint + viewer; real v1→v2 diffs verified"),
 ("REG-503","E7",5,"Rich document drawer",
  "Document view with tabs: rendered markdown (self-contained md→html, no CDN), summary, provenance rail (all governance fields incl. final URL/fetch method/content source), raw text.",
  "Headings/lists/tables/links render; provenance shows source URL clickable; raw preserved verbatim",
  3,"Full-stack",D,"md2html engine; verified on live BoE paper"),
 ("REG-504","E7",5,"Corpus file Explorer",
  "Nested file-tree page mirroring disk (regulator→current|archive→doc_type→year→doc→artifacts) with lazy loading; text files open in drawer, originals download; API jailed to corpus root.",
  "Path traversal rejected (400); item counts and sizes shown; md renders rich",
  3,"Full-stack",D,"#exp; jail tests"),
 ("REG-505","E5",5,"Manual interjection & upload lane",
  "Human override: add/update a document by URL fetch, pasted corrected markdown, or uploaded PDF/HTML — via UI modal, API, and CLI — through the identical versioning/provenance path, tagged manual and audited.",
  "Update archives prior version; operator identity recorded; run row visible; future crawls recognize the URL (no duplicates); 50MB upload cap",
  5,"Backend",D,"manual.py + modal + /documents(/upload) + CLI; v1→v2 test"),
 ("REG-506","E8",5,"Content-plane MCP tools",
  "reg_search (filter-then-rank), reg_read (summary|full|meta|historical version), reg_tags, reg_whats_new, reg_graph — stateless, registered via config-driven dynamic discovery.",
  "All five discoverable and callable over MCP; filters applied before ranking; graph traversal cycle-safe depth≤3",
  5,"AI",D,"reg_tools.py + generated configs; tool tests"),
 ("REG-507","E8",5,"Index-plane MCP tools",
  "reg_coverage, reg_browse (metadata-only), reg_changes, reg_diff, reg_inventory, reg_runs_status, reg_trigger_run (mutating, guarded) — the answers content-RAG cannot give.",
  "All seven callable over MCP with the same filters as the UI; trigger refuses concurrent runs and records operator",
  5,"AI",D,"Verified over the wire via agent key"),
 ("REG-508","E8",5,"MCP security & key lifecycle",
  "Authentication required for tools/call (discovery stays open); per-key tool allowlists enforced at the transport; scoped agent key issued, rotated, and stored DB-only (never committed).",
  "No-key and dead-key calls rejected; out-of-scope tool denied; in-scope tool works — four-way matrix verified; rotation procedure documented",
  5,"Backend",D,"api_routes enforcement; live verification matrix; key rotated pre-publication"),
 ("REG-509","E7",5,"BA: UAT round 2 (changes/diff/explorer/manual)",
  "Structured UAT of the remaining pages and the manual lane.",
  "UAT script executed; defects triaged; sign-off recorded",
  2,"BA",PR,"In-build PO validation only; formal pass pending"),
 ("PM-S5","E5",5,"PM: sprint ceremonies & reporting","Recurring PM allocation.","As S1",0,"PM",D,"Recurring"),

 # Sprint 6 — Agentic integration, ops close-out, docs
 ("REG-601","E9",6,"Markdown projection for the agent stack",
  "Maintain the corpus in the agent tools' consumption layout — data/markdown/{web|policy}/{regulator}/{doc_type}/{doc_id}.md with YAML frontmatter — via write-through on every ingest plus nightly resync self-heal.",
  "Every current document has exactly one projection file; frontmatter carries citation context; resync idempotent; history stays in canonical store",
  5,"Backend",D,"projection.py; 5,957 files (269MB); write-through test"),
 ("REG-602","E9",6,"Agent-platform MCP connectivity",
  "Connect the existing agent stack to the MCP endpoint with the scoped key; verify end-to-end tool calls and denial of out-of-scope tools.",
  "initialize/tools list/call verified from client side; connection details documented for the platform's custom-MCP dialog",
  3,"AI",D,"Verified (usage counter); key details delivered"),
 ("REG-603","E9",6,"Chatbot enablement guide",
  "Author the integration guide: connection block, agent prompt pattern (narrow→read cheap→relate→freshness), non-negotiable citation rule, worked examples, honest limitations, go-live eval checklist.",
  "Guide sufficient for the chatbot team to integrate without our involvement",
  2,"AI",D,"docs/regagg/CHATBOT_GUIDE.md"),
 ("REG-604","E9",6,"Agent skills authoring",
  "Author skill definitions on the agent platform mapping to the tools: Daily Regulatory Briefing (reg_changes), Impact Analysis (reg_diff+reg_graph), Completeness Audit (reg_coverage+reg_inventory).",
  "Three skills live on the platform; each produces correct, cited output on test prompts",
  3,"AI",NS,"Owner's agent platform; guide + tool surface ready"),
 ("REG-605","E9",6,"Agent citation & accuracy evaluation",
  "Run the 20-question eval across ≥8 regulators: every claim cited (regulator+reference+date+URL), superseded docs flagged with successor, out-of-corpus questions declined.",
  "Eval sheet completed with pass/fail per question; failures fed back into prompt/tool fixes",
  3,"AI",NS,"Checklist ready in CHATBOT_GUIDE; requires live agent stack"),
 ("REG-606","E5",6,"Daily delta poller & weekly deep run",
  "Single scheduler entrypoint: fleet delta ingest (giants depth-sliced), pending-edge resolution, integrity reconcile, projection resync, one-line delta summary; weekly --deep raises caps.",
  "One cron line each for daily/weekly; exit code signals invariant violations; every run visible in UI",
  3,"Backend",D,"regagg_daily_poll.py; cron lines documented"),
 ("REG-607","E5",6,"Backup & restore",
  "Nightly consistent SQLite snapshot + hardlink-incremental corpus backup with 7-day retention; restore verified.",
  "Backup runs while server/crawls active; restore drill opens snapshot with full document count; retention enforced portably",
  3,"Backend",D,"regagg_backup.sh; 2.0GB verified restore-check"),
 ("REG-608","E7",6,"UX polish round",
  "Excerpt/description line + source links + Added date in corpus; PDF filename titles (incl. backfill); add-doc discoverability in top bar; rich-render defaults.",
  "No raw URLs as titles; every row shows an honest excerpt where text exists; date column always populated",
  3,"Full-stack",D,"729 titles backfilled; excerpts live"),
 ("REG-609","E10",6,"Successor knowledge base",
  "Author the handover KB: start-here mental model, architecture/flows/schema, UI+tools reference, testing/ops runbook, known issues + war stories + roadmap.",
  "A new engineer/AI agent can operate and extend the system from the KB alone",
  3,"Full-stack",D,"docs/regagg/kb/ (5 docs)"),
 ("REG-610","E9",6,"BA: final UAT & product-owner acceptance",
  "End-to-end acceptance against the epic-level outcomes; PO sign-off for go-live.",
  "Acceptance record signed; open items dispositioned to backlog",
  2,"BA",NS,"Pending PO session"),
 ("PM-S6","E5",6,"PM: go/no-go readiness review & handover","Deployment readiness review, handover of runbooks/KB, closure report.","Readiness review held; handover acknowledged",0,"PM",NS,"Pending"),

 # Backlog — E10 future enhancements
 ("REG-701","E10","Backlog","Federal Register full text via govinfo.gov bulk data",
  "Replace abstract/stub fallback with sanctioned full text from GPO govinfo bulk repository (separate host built for programmatic bulk use).",
  "Full text for new FR docs; existing stubs self-heal to full text as revisions; no ToS risk",
  5,"AI",NS,"Roadmap #1; adapter pattern ready"),
 ("REG-702","E10","Backlog","Simple-mode overview landing page",
  "First-time-user landing: one plain-English headline, ≤4 tiles, one regulator×(web/PDF/today delta/health-dot) table; ops detail moves behind drill-ins; analyst vs operator modes.",
  "New user answers 'what do we have / what changed today / is it healthy' in <10s without training",
  5,"Full-stack",NS,"UX proposal documented"),
 ("REG-703","E10","Backlog","Playwright fetch path for JS-heavy sites",
  "Browser-rendered fetching for regulators whose listings require JavaScript (MAS, some NYDFS).",
  "fetch:playwright configs ingest content where plain HTTP yields none",
  5,"AI",NS,"Config field already parses"),
 ("REG-704","E10","Backlog","PostgreSQL migration",
  "Flip db.type to Postgres using the shipped DDL; enable pg_trgm for title-similarity reference resolution.",
  "All tests green on Postgres; no write-lock contention during concurrent crawl+agent load",
  3,"Backend",NS,"DDL ready"),
 ("REG-705","E10","Backlog","Expected inventories for EBA / FCA / APRA",
  "Extend completeness proofs beyond OSFI using each regulator's official register/index.",
  "Per-series present/missing for 3 more regulators; gaps drive targeted gap-fills",
  3,"AI",NS,""),
 ("REG-706","E10","Backlog","Weekly digest webhook",
  "Teams/email digest generated from the changes feed (per-desk filters).",
  "Scheduled digest delivered with citations; opt-in filters per audience",
  3,"Backend",NS,"Spec FR-9"),
 ("REG-707","E10","Backlog","AMF Québec access via official whitelisting",
  "Engage AMF for institutional IP/UA whitelisting (site bot-blocks all automation); fallback: licensed feed.",
  "AMF collecting via sanctioned access; zero evasion techniques used",
  2,"BA",NS,"Escalated; contact pending"),
 ("REG-708","E10","Backlog","Production hardening",
  "Pre-exposure hardening: replace default admin credentials, TLS/reverse proxy, restricted bind, secrets management, key rotation SOP.",
  "Security checklist passed before any non-localhost exposure",
  5,"Backend",NS,"MANDATORY before deployment beyond localhost"),
 ("REG-709","E10","Backlog","OCR for scanned policy PDFs",
  "Tesseract path for image-only PDFs (currently 6 docs, 97% extract fine without).",
  "ocr:true docs gain searchable text; volume-triggered (revisit if scanned share grows)",
  3,"AI",NS,"Deprioritized by owner 2026-08-04"),
 ("REG-710","E10","Backlog","LLM enrichment integration",
  "Wire the owner's LLM gateway into the enrichment slot: real summaries (summary.md), taxonomy topic tags, doc-type reclassification, date/deadline extraction; golden-set eval gates.",
  "Summaries replace excerpts; doc_type distribution reflects reality (not 90% announcement); deadline feed populated; eval ≥90% tagging accuracy",
  8,"AI",NS,"MockLLM slot + eval harness pattern in place"),
]

ROLES = {"Backend": "Backend Engineer", "Full-stack": "Full-stack Engineer",
         "AI": "AI Engineer", "BA": "Business Analyst", "PM": "Project Manager"}

# ── test-case inventory ─────────────────────────────────────────────────────
FILE_MAP = {  # test file -> (story, epic, type, what it verifies)
 "test_versioning.py": ("REG-201/202/203","E2","Unit+Chaos","Atomic override→archive protocol; crash injected at every step; reconcile restores invariants with zero loss"),
 "test_connectors_pipeline.py": ("REG-204/205/206/301/306","E3","Integration","Connector detection per engine vs real configs; OSFI end-to-end incl. forced update→archive; rerun idempotency"),
 "test_orchestrator.py": ("REG-304","E5","Integration","Fleet fan-out failure isolation; audited subset rerun; reconcile"),
 "test_enrichment.py": ("REG-710","E6","Unit","LLM-slot enrichment contract: taxonomy tags, dates never guessed, supersedes flip, enrichment_pending fallback"),
 "test_rules_and_capture.py": ("REG-405/406/302/303/601","E6","Unit+Integration","Reference grammars; sentence-scoped supersede; PDF harvest + source_kind; meta-source dedup; backfill cutoff; markdown projection"),
 "test_mcp_tools.py": ("REG-506/507","E8","Integration","All 12 MCP tools against seeded corpus; generated configs; trigger delegates to runtime"),
 "test_ui_endpoints.py": ("REG-401/403/501/502/504/505","E7","Integration+Regression","Tree/browse/changes/diff/runs/inventory endpoints; filter-before-LIMIT regression; manual add/update; multipart upload; fs jail"),
 "test_admin.py": ("REG-401/402","E7","Integration","Coverage matrix, drill-down, audited rerun/toggle, integrity"),
 "test_fetch.py": ("REG-207","E3","Unit+Regression","HTML→md; PDF magic-byte routing (block-page regression); URL-title humanizer"),
 "test_verify_sources.py": ("REG-107","E4","Unit","Reachability/content-type/parse/freshness checks; sitemap-index; report rendering"),
 "test_recorded_fixtures.py": ("REG-108","E4","E2E-Fixture","Every regulator's connector parses its real recorded payload (29 regs)"),
}

MANUAL_CASES = [
 ("MTC-01","REG-508","E8","Security-Manual","MCP auth matrix over the wire","No key→401; disabled key→401; valid key + reg tool→200; valid key + out-of-scope tool→denied (-32001)","Pass","curl matrix 2026-08-04"),
 ("MTC-02","REG-508","E8","Security-Manual","API key rotation","Old key dead everywhere; new DB-only key works; committed copy scrubbed before repo publication","Pass","rotation 2026-08-03"),
 ("MTC-03","REG-401","E7","Security-Manual","Connection-pool endurance","30 rapid API calls with 5s-polling load; no pool exhaustion","Pass","hammer test post thread-scoped sessions"),
 ("MTC-04","REG-407","E6","System","OSFI completeness reconciliation","Inventory reports 30/30 (CAR 9/9, LAR 5/5, B 8/8, E 8/8) after guidance gap-fill","Pass","/inventory/osfi"),
 ("MTC-05","REG-307/408","E4","System","Live fleet collection","29/30 regulators verified and collecting; ~5,970 docs; per-regulator coverage % reported","Pass","live crawls 2026-08-02..04"),
 ("MTC-06","REG-309","E5","System","FedReg fallback chain","All ~500 FR docs carry abstract or labeled stub; content_source in provenance; zero error storm","Pass","fedreg re-ingests"),
 ("MTC-07","REG-607","E5","System","Backup + restore check","Consistent snapshot while live; restored DB opens with full doc count; retention pruning","Pass","2.0GB backup restore-check"),
 ("MTC-08","REG-201","E2","System","Live update→archive","Real revised pages produce v2 with prior version archived and diff viewable","Pass","6 live revisions (frb/boe_pra/fca)"),
 ("MTC-09","REG-402..404","E7","UAT-Manual","Dashboard smoke on live data","All 5 pages render live corpus; drill-downs, drawer, diff, rerun guard behave","Pass","browser walkthroughs"),
 ("MTC-10","REG-505","E5","UAT-Manual","Manual lane end-to-end","URL-fetch, pasted-markdown, and PDF-upload paths create/version docs with manual tag + audit","Pass","API+UI verification"),
 ("MTC-11","REG-601","E9","System","Projection integrity","5,957 projection files match current corpus; frontmatter valid; resync idempotent","Pass","resync report"),
 ("MTC-12","REG-304","E5","System","Concurrent-run guard","Second fleet trigger refused while a run is active","Pass","live rerun attempt"),
 ("MTC-13","REG-106","E1","Gate","Foundation gate","30/30 configs parse; taxonomy sync; 9 tables register/create","Pass","regagg_verify_foundation.py"),
 ("MTC-14","REG-605","E9","UAT-Manual","Agent citation eval (20 questions)","Every answer cites regulator+reference+date+URL; declines out-of-corpus","Not Run","awaiting owner agent stack"),
]

def load_test_ids():
    p = Path("/tmp/test_ids.txt")
    ids = []
    if p.exists():
        for line in p.read_text().splitlines():
            line = line.strip()
            if "::" in line:
                f = line.split("::")[0].split("/")[-1]
                ids.append((f, line.split("::", 1)[1].split("[")[0], line))
    return ids


# ── CSV writers ─────────────────────────────────────────────────────────────
PLAN_COLS = ["Story ID","Epic","Epic Name","Sprint","Title","Description",
             "Acceptance Criteria","Story Points","Assigned Role"]
EXEC_COLS = PLAN_COLS + ["Status","Evidence / Notes"]
EPIC_BY = {e[0]: e[1] for e in EPICS}

def story_row(s, with_status):
    sid, epic, sprint, title, desc, ac, pts, role, status, note = s
    base = [sid, epic, EPIC_BY[epic],
            f"Sprint {sprint}" if isinstance(sprint, int) else sprint,
            title, desc, ac, pts, ROLES[role]]
    return base + [status, note] if with_status else base

with open(PLAN / "sprint_plan.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(PLAN_COLS)
    for s in S: w.writerow(story_row(s, False))

with open(PLAN / "epics.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(["Epic","Name","Objective"])
    for e in EPICS: w.writerow(e)

CAP = [["Role","Allocation","Days / Sprint","SP Capacity / Sprint","Sprints","Notes"],
       ["Backend Engineer","100%",10,11,6,"Data layer, versioning, pipeline, ops"],
       ["Full-stack Engineer","100%",10,11,6,"Dashboard UI, fetch layer, docs"],
       ["AI Engineer","100%",10,11,6,"Connectors, enrichment, MCP tools, agentic"],
       ["Business Analyst","25%",2.5,2,6,"Source sign-off, UX definitions, UAT"],
       ["Project Manager","25%",2.5,0,6,"Ceremonies, RAID, reporting (non-pointed)"]]
with open(PLAN / "capacity.csv", "w", newline="") as f:
    csv.writer(f).writerows(CAP)

with open(EXEC / "story_status.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(EXEC_COLS)
    for s in S: w.writerow(story_row(s, True))

TC_COLS = ["TC ID","Story ID","Epic","Type","Test Case","Verifies","Status","Evidence"]
tc_rows = []
for i, (fname, tname, full) in enumerate(load_test_ids(), 1):
    story, epic, typ, what = FILE_MAP.get(fname, ("-","-","Unit","-"))
    tc_rows.append([f"ATC-{i:02d}", story, epic, typ, tname, what, "Pass", full])
tc_rows += [list(r) for r in MANUAL_CASES]
with open(EXEC / "test_cases.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(TC_COLS); w.writerows(tc_rows)

# ── Excel workbooks ─────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

def sheet_from_rows(ws, header, rows, widths):
    ws.append(header)
    for c in ws[1]:
        c.fill, c.font = HDR_FILL, HDR_FONT
    for r in rows:
        ws.append(r)
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font, c.alignment = BODY, WRAP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# planning workbook
wb = Workbook(); wb.remove(wb.active)
sheet_from_rows(wb.create_sheet("Roadmap"), PLAN_COLS,
                [story_row(s, False) for s in S],
                [10,6,30,9,34,52,52,7,17])
sheet_from_rows(wb.create_sheet("Epics"), ["Epic","Name","Objective"],
                [list(e) for e in EPICS], [7,38,90])
cap_ws = wb.create_sheet("Capacity")
sheet_from_rows(cap_ws, CAP[0], CAP[1:], [20,11,14,20,9,44])
cap_ws["A8"] = "Planned SP by sprint (dev+BA):"; cap_ws["A8"].font = Font(name="Arial", bold=True, size=10)
for i in range(1, 7):
    cap_ws.cell(row=9, column=i, value=f"Sprint {i}").font = HDR_FONT
    cap_ws.cell(row=9, column=i).fill = HDR_FILL
    cap_ws.cell(row=10, column=i,
                value=f'=SUMIFS(Roadmap!H:H,Roadmap!D:D,"Sprint {i}")').font = BODY
cap_ws.cell(row=9, column=7, value="Backlog").font = HDR_FONT
cap_ws.cell(row=9, column=7).fill = HDR_FILL
cap_ws.cell(row=10, column=7, value='=SUMIFS(Roadmap!H:H,Roadmap!D:D,"Backlog")').font = BODY
wb.save(PLAN / "Regulatory_Aggregator_Roadmap.xlsx")

# execution workbook
wb2 = Workbook(); wb2.remove(wb2.active)
sheet_from_rows(wb2.create_sheet("Story Status"), EXEC_COLS,
                [story_row(s, True) for s in S],
                [10,6,30,9,34,50,50,7,17,11,40])
sheet_from_rows(wb2.create_sheet("Test Cases"), TC_COLS, tc_rows,
                [9,18,7,17,40,60,9,44])
sm = wb2.create_sheet("Summary")
sm["A1"] = "Execution Summary"; sm["A1"].font = Font(name="Arial", bold=True, size=12)
rows = [("Total stories", "=COUNTA('Story Status'!A:A)-1"),
        ("Done", "=COUNTIF('Story Status'!J:J,\"Done\")"),
        ("Partial", "=COUNTIF('Story Status'!J:J,\"Partial\")"),
        ("Not started", "=COUNTIF('Story Status'!J:J,\"Not Started\")"),
        ("Story points — Done", "=SUMIFS('Story Status'!H:H,'Story Status'!J:J,\"Done\")"),
        ("Story points — total (incl. backlog)", "=SUM('Story Status'!H:H)"),
        ("Automated test cases (Pass)", "=COUNTIFS('Test Cases'!D:D,\"<>*Manual*\",'Test Cases'!G:G,\"Pass\")"),
        ("Manual/system verifications (Pass)", "=COUNTIFS('Test Cases'!D:D,\"*Manual*\",'Test Cases'!G:G,\"Pass\")+COUNTIFS('Test Cases'!D:D,\"System\",'Test Cases'!G:G,\"Pass\")+COUNTIFS('Test Cases'!D:D,\"Gate\",'Test Cases'!G:G,\"Pass\")"),
        ("Pending verifications", "=COUNTIF('Test Cases'!G:G,\"Not Run\")")]
for i, (label, formula) in enumerate(rows, 3):
    sm.cell(row=i, column=1, value=label).font = BODY
    sm.cell(row=i, column=2, value=formula).font = Font(name="Arial", bold=True, size=10)
sm.column_dimensions["A"].width = 38; sm.column_dimensions["B"].width = 14
wb2.save(EXEC / "Execution_Status.xlsx")

dev_sp = {}
for s in S:
    key = f"S{s[2]}" if isinstance(s[2], int) else "Backlog"
    dev_sp[key] = dev_sp.get(key, 0) + s[6]
print("SP per sprint:", dict(sorted(dev_sp.items())))
print(f"stories: {len(S)} · test rows: {len(tc_rows)}")
print("written to", PLAN, "and", EXEC)
